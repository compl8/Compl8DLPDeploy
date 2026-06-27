import importlib.util, os, sys
from pathlib import Path

SCRIPTS = Path(__file__).resolve().parent.parent / "scripts"
sys.path.insert(0, str(SCRIPTS))

def _load_cli():
    spec = importlib.util.spec_from_file_location("sync_spreadsheet", SCRIPTS / "sync-spreadsheet.py")
    mod = importlib.util.module_from_spec(spec); spec.loader.exec_module(mod)
    return mod

def _make_xlsx(path, rows):
    import openpyxl, catalog_refresh as cr
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "SIT Risk Analysis"
    header = [""] * cr.SHEET_WIDTH
    header[cr.COL_NAME] = "SIT Name"; header[cr.COL_SLUG] = "GUID / Slug"; header[cr.COL_SOURCE] = "Source"
    ws.append(header)
    for r in rows: ws.append(r)
    wb.save(str(path)); wb.close()

def test_build_change_report(tmp_path):
    import catalog_refresh as cr
    cli = _load_cli()
    rows = [cr.row_from_catalog({"slug": "keep", "name": "Keep", "version": "1.0.0"}),
            cr.row_from_catalog({"slug": "gone", "name": "Gone", "version": "1.0.0"})]
    catalog = [{"slug": "keep", "name": "Keep", "version": "1.2.0"},   # tuned
               {"slug": "new", "name": "New", "version": "1.0.0"},      # added
               {"slug": "msft", "name": "Ms", "source": "microsoft"}]   # skipped
    rep = cli.build_change_report(catalog, rows)
    assert rep["added"] == ["new"]
    assert rep["removed"] == ["gone"]
    assert rep["tuned"] == ["keep"]
    assert rep["skipped_microsoft"] == 1

def test_write_report_emits_files(tmp_path):
    cli = _load_cli()
    rep = {"added": ["a"], "removed": ["b"], "drift": [("c", ["name"])], "tuned": ["d"], "skipped_microsoft": 2}
    cli.write_report(tmp_path, rep)
    assert (tmp_path / "changes.csv").exists()
    md = (tmp_path / "report.md").read_text(encoding="utf-8")
    assert "Added: 1" in md and "Tuned" in md

def test_apply_update_appends_and_preserves(tmp_path):
    import openpyxl, catalog_refresh as cr
    from catalog_update import apply_update
    # existing row with a human-curated risk_rating + tier flag we must preserve
    existing = cr.row_from_catalog({"slug": "keep", "name": "Keep", "version": "1.0.0"})
    existing[cr.COL_RISK_RATING] = 7              # curated
    existing[9] = "Y"                             # Small (tenant) tier flag (curated)
    src = tmp_path / "in.xlsx"
    _make_xlsx(src, [existing])
    catalog = {"keep": {"slug": "keep", "name": "Keep", "version": "1.3.0"},  # tuned -> sync version
               "new": {"slug": "new", "name": "New", "data_categories": ["pii"],
                       "version": "1.0.0", "source": None}}
    out = tmp_path / "out.xlsx"
    apply_update(src, out, catalog, added=["new"], removed=[], refresh_metadata=False, in_place=False)

    wb = openpyxl.load_workbook(str(out)); ws = wb["SIT Risk Analysis"]
    data = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]; wb.close()
    by_slug = {r[cr.COL_SLUG]: r for r in data}
    assert set(by_slug) == {"keep", "new"}                      # appended
    assert by_slug["keep"][cr.COL_RISK_RATING] == 7            # curated preserved
    assert by_slug["keep"][9] == "Y"                           # tier flag preserved
    assert str(by_slug["keep"][cr.COL_VERSION]) == "1.3.0"     # freshness synced
    assert by_slug["new"][cr.COL_CATEGORY] == "pii"            # new row enriched
    assert src.exists()                                        # input untouched

def test_apply_update_idempotent(tmp_path):
    import openpyxl, catalog_refresh as cr
    from catalog_update import apply_update
    existing = cr.row_from_catalog({"slug": "keep", "name": "Keep", "version": "1.0.0"})
    src = tmp_path / "in.xlsx"; _make_xlsx(src, [existing])
    catalog = {"keep": {"slug": "keep", "name": "Keep", "version": "1.0.0"}}
    out = tmp_path / "out.xlsx"
    apply_update(src, out, catalog, added=[], removed=[], refresh_metadata=False, in_place=False)
    wb = openpyxl.load_workbook(str(out)); ws = wb["SIT Risk Analysis"]
    n = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[cr.COL_SLUG]); wb.close()
    assert n == 1   # no duplicate appended
