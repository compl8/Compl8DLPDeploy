"""openpyxl write path for the catalog refresh: append new rows + sync freshness on
existing rows, preserving curated values. Side-effectful (separated from the pure
catalog_refresh module). See docs/superpowers/specs/2026-06-27-catalog-refresh-design.md.
"""
import shutil
from pathlib import Path

import catalog_refresh as cr

SHEET = "SIT Risk Analysis"

# Catalog-sourced columns refreshed on existing rows ONLY under --refresh-metadata.
# Excludes risk_rating + all curated columns by construction.
_META_REFRESH = (
    (cr.COL_NAME, lambda p: str(p.get("name") or "")),
    (cr.COL_CATEGORY, lambda p: cr._join(p.get("data_categories"))),
    (cr.COL_JURISDICTIONS, lambda p: cr._join(p.get("jurisdictions"))),
    (cr.COL_CONFIDENCE, lambda p: str(p.get("confidence") or "")),
    (cr.COL_REGULATIONS, lambda p: cr._join(p.get("regulations"))),
    (cr.COL_RISK_DESC, lambda p: str(p.get("risk_description") or "")),
)
# Freshness provenance: synced on every run regardless of --refresh-metadata.
_FRESHNESS = (
    (cr.COL_CREATED, lambda p: str(p.get("created") or "")),
    (cr.COL_UPDATED, lambda p: str(p.get("updated") or "")),
    (cr.COL_VERSION, lambda p: str(p.get("version") or "")),
)


def apply_update(xls_path, out_path, catalog, added, removed, refresh_metadata, in_place):
    import openpyxl
    if in_place:
        shutil.copy2(str(xls_path), str(xls_path) + ".bak")
        out_path = xls_path
    wb = openpyxl.load_workbook(str(xls_path))   # keep formulas + other sheets
    ws = wb[SHEET]

    slug_to_row = {}
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        name = row[cr.COL_NAME].value
        slug = row[cr.COL_SLUG].value
        if name and slug:
            slug_to_row[str(slug).strip()] = i

    for s in added:
        ws.append(cr.row_from_catalog(catalog[s]))

    for s, excel_row in slug_to_row.items():
        p = catalog.get(s)
        if not p:
            continue
        for col, getter in _FRESHNESS:
            ws.cell(row=excel_row, column=col + 1, value=getter(p))
        if refresh_metadata:
            for col, getter in _META_REFRESH:
                ws.cell(row=excel_row, column=col + 1, value=getter(p))

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    wb.close()
