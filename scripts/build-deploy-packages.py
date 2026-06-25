#!/usr/bin/env python3
"""Build deployment packages by fetching right-sized batches from testpattern.dev.

Fetches entities in batches from testpattern.dev's purview-bundle API with
dictionary placeholders, strips whitespace, and writes directly as UTF-8.
No XML parsing or reassembly — preserves the API's valid XML exactly.

If a batch exceeds 150KB, it's split in half and re-fetched.

Usage:
  python scripts/build-deploy-packages.py --tier medium
  python scripts/build-deploy-packages.py --tier medium --max-terms 5
  python scripts/build-deploy-packages.py --tier medium --dry-run
"""

import argparse
import json
import os
import re
import sys
import time
import urllib.request
import urllib.error
from pathlib import Path

MAX_PACKAGE_UTF16 = 150 * 1024        # 153600 — Purview/testpattern hard cap, measured UTF-16LE
PREFERRED_PACKAGE_UTF16 = 148 * 1024  # 151552 — working budget under the cap
MAX_PACKAGES = 9                      # automated build cap; 10th tenant slot reserved for manual adds
MAX_SITS_PER_PACKAGE = 50             # Purview authoring hard cap: 50 entities per rule package
TESTPATTERN_API = "https://testpattern.dev/api/export/purview-bundle"
TIER_COLS = {"small": 9, "medium": 10, "large": 11}


def utf16_len(text):
    return len(text.encode("utf-16-le"))


def select_bin(remaining, sizes, shrink, target, max_entities=MAX_SITS_PER_PACKAGE):
    """Greedily choose slugs (caller passes `remaining` sorted desc by sizes[s]) whose estimated
    real size (single-slug size * shrink) sums to ~target. Pure. Returns (chosen, rest).
    Also enforces a hard entity cap: once len(chosen) >= max_entities remaining slugs go to rest."""
    chosen, rest, est = [], [], 0.0
    for s in remaining:
        if len(chosen) >= max_entities:
            rest.append(s)
            continue
        e = sizes[s] * shrink
        if est + e <= target:
            chosen.append(s); est += e
        else:
            rest.append(s)
    return chosen, rest


def load_slugs(xls_path, tier):
    """Load TestPattern slugs for the requested tier."""
    import openpyxl
    tier_col = TIER_COLS[tier]
    wb = openpyxl.load_workbook(str(xls_path), read_only=True, data_only=True)
    ws = wb["SIT Risk Analysis"]
    slugs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        slug = str(row[1] or "").strip()
        source = str(row[14] or "").strip()
        in_tier = str(row[TIER_COLS[tier]] or "").strip().upper() == "Y"
        is_uuid = (
            len(slug) == 36
            and slug.count("-") == 4
            and all(c in "0123456789abcdef-" for c in slug.lower())
        )
        if in_tier and source == "TestPattern" and slug and not is_uuid:
            slugs.append(slug)
    wb.close()
    return slugs


def load_settings(project_root):
    """Load prefix and publisher from settings.json."""
    path = os.path.join(str(project_root), "config", "settings.json")
    if os.path.isfile(path):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        return data.get("namingPrefix", "DLP"), data.get("publisher", "")
    return "DLP", ""


def fetch_bundle(slugs, name):
    """Fetch a purview-bundle XML from testpattern.dev with dictionaries."""
    slug_param = ",".join(slugs)
    url = f"{TESTPATTERN_API}?slugs={slug_param}&name={name}&dictionaries=true"
    req = urllib.request.Request(url, headers={"User-Agent": "Compl8DLPDeploy/1.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        raw = resp.read()
    # The API serves UTF-16 LE with BOM (testpattern's native encoding); older responses
    # were UTF-8. The toolkit's recorded encoding decision is UTF-8, so transcode here and
    # normalise the declaration — the "modulo declared encoding decision" of the design's
    # Stage 3 parity rule.
    if raw.startswith(b"\xff\xfe"):
        text = raw.decode("utf-16-le").lstrip("﻿")
    else:
        text = raw.decode("utf-8")
    return text.replace(
        '<?xml version="1.0" encoding="utf-16"?>',
        '<?xml version="1.0" encoding="utf-8"?>',
        1,
    )


def optimise(xml_text, publisher):
    """Strip whitespace/comments and patch publisher. Preserves raw XML structure."""
    xml_text = re.sub(r"<!--.*?-->", "", xml_text, flags=re.DOTALL)
    xml_text = re.sub(r">\s+<", ">\n<", xml_text)
    if publisher:
        xml_text = re.sub(
            r"<PublisherName>[^<]*</PublisherName>",
            f"<PublisherName>{publisher}</PublisherName>",
            xml_text,
        )
    return xml_text


def count_entities(xml_text):
    return len(re.findall(r'<Entity\s+id="', xml_text))


def measure_slugs(slugs, single, measure_fn, delay=0.0, attempts=3, sleep_fn=time.sleep, log=print):
    """Measure each slug's footprint into `single` (slug -> size), retrying transient failures.

    A single-slug request can fail transiently (timeout/429/5xx). Such a slug must NOT be silently
    dropped from the build — it is retried up to `attempts` times across sweeps. `measure_fn(slug)`
    returns the size (raising on failure). Returns (single, missing) where `missing` is the slugs
    still unmeasured after all attempts; the caller decides (we fail the build) — never write an
    incomplete package set. `single` is mutated/returned so a persisted cache lets a re-run resume.
    """
    to_measure = [s for s in slugs if not single.get(s)]
    for attempt in range(1, attempts + 1):
        pending = [s for s in to_measure if not single.get(s)]
        if not pending:
            break
        if attempt > 1:
            log(f"  Retry {attempt - 1}/{attempts - 1}: re-measuring {len(pending)} slug(s) that failed...")
        for s in pending:
            try:
                single[s] = measure_fn(s)
            except Exception as e:
                log(f"    measure {s}: FAILED {e}")
            sleep_fn(delay)
    missing = [s for s in slugs if not single.get(s)]
    return single, missing


def main():
    parser = argparse.ArgumentParser(description="Build deployment packages from testpattern.dev")
    parser.add_argument("--tier", choices=["small", "medium", "large"], default="medium")
    parser.add_argument("--xls", default=None)
    parser.add_argument("--output-dir", default="xml/deploy")
    parser.add_argument("--batch-size", type=int, default=35, help="Slugs per API call")
    parser.add_argument("--prefix", default=None, help="Package name prefix (default: from settings.json)")
    parser.add_argument("--publisher", default=None, help="Publisher name (default: from settings.json)")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--delay", type=float, default=1.0)
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parent.parent
    output_dir = str((project_root / args.output_dir).resolve())

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from pipeline_utils import find_spreadsheet
    xls_path = find_spreadsheet(project_root, args.xls)
    if not xls_path:
        print("ERROR: No input spreadsheet found. Specify with --xls or set inputSpreadsheet in settings.json", file=sys.stderr)
        sys.exit(1)

    cfg_prefix, cfg_publisher = load_settings(project_root)
    prefix = args.prefix or cfg_prefix
    publisher = args.publisher or cfg_publisher

    print(f"  Source: {xls_path.name}")
    print(f"  Tier: {args.tier}")
    print(f"  Batch size: {args.batch_size}")
    print(f"  Prefix: {prefix}")

    # Load slugs
    slugs = load_slugs(xls_path, args.tier)
    print(f"  Slugs: {len(slugs)}")

    # Exclude verified Microsoft built-ins from PACKING — they're referenced by GUID in the DLP
    # rules (Build-FromXLS), already exist in the tenant, and don't count toward our package limit.
    from _bfx_loaders import load_builtin_map
    builtin_map = load_builtin_map(os.path.join(str(project_root), "tenant-sits-full.csv"))
    if builtin_map:
        import openpyxl
        _wb = openpyxl.load_workbook(str(xls_path), read_only=True, data_only=True)
        _ws = _wb["SIT Risk Analysis"]
        name_by_slug = {str(r[1]).strip(): str(r[0]).strip()
                        for r in _ws.iter_rows(min_row=2, values_only=True) if r[0] and r[1]}
        _wb.close()
        _norm = lambda n: "".join(c for c in str(n or "").lower() if c.isalnum())
        before = len(slugs)
        slugs = [s for s in slugs if _norm(name_by_slug.get(s, "")) not in builtin_map]
        print(f"  Excluded {before - len(slugs)} Microsoft built-ins from packing (referenced by GUID instead)")

    if args.dry_run:
        n = (len(slugs) + args.batch_size - 1) // args.batch_size
        print(f"\n  [DRY RUN] Would create ~{n} packages")
        return

    os.makedirs(output_dir, exist_ok=True)

    # ---- Phase 1: measure each slug's UTF-16 footprint (single-slug bundle), cached ----
    cache_path = os.path.join(output_dir, ".size-cache.json")
    single = {}
    if os.path.isfile(cache_path):
        with open(cache_path, encoding="utf-8") as f:
            single = json.load(f)
    to_measure = [s for s in slugs if not single.get(s)]
    print(f"  Measuring {len(to_measure)} slugs (cached: {len(slugs) - len(to_measure)})...")
    single, missing = measure_slugs(
        slugs, single,
        lambda s: utf16_len(optimise(fetch_bundle([s], "measure"), publisher)),
        delay=args.delay,
    )
    # Persist whatever we measured (even on partial failure) so a re-run resumes from the cache.
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump({k: v for k, v in single.items() if v}, f, indent=2)
    if missing:
        # A measurement failure means we cannot know a slug's size, so it would be silently absent
        # from every package. Refuse to write an incomplete build — fail loudly; the operator re-runs
        # (cached measurements make the retry cheap) or fixes connectivity.
        print(f"  ERROR: {len(missing)} slug(s) could not be measured after retries: {missing}", file=sys.stderr)
        print("  Refusing to write an incomplete package set. Re-run to retry (measurements are cached).", file=sys.stderr)
        sys.exit(1)
    live = {s: single[s] for s in slugs if single.get(s)}

    # ---- Phase 2+3: ITERATIVE MEASURED bin-packing (fetch-verify, self-calibrating) ----
    # Static estimation can't reliably hit the tight ~145KB target (the real/single shrink varies
    # per slug set). Instead: pick a bin by ESTIMATE (single-slug size * shrink ~= target), FETCH it,
    # measure the REAL size, back off the largest slug while it's over ACCEPT, then re-calibrate
    # `shrink` from the real bin so the next estimate is sharper. Greedy largest-first; <=9 packages.
    TARGET = 145 * 1024            # estimated real-size fill target per package
    ACCEPT = PREFERRED_PACKAGE_UTF16  # 151552 — accept a fetched bin at/under this (under the 153600 hard cap)
    shrink = 0.84                  # initial real/single ratio; refined after each real bin
    remaining = sorted(live, key=lambda s: -live[s])
    results, dropped = [], []
    pkgnum = 0
    while remaining and pkgnum < MAX_PACKAGES:
        pkgnum += 1
        pkg_name = f"{prefix}-{args.tier}-{pkgnum:02d}"
        chosen, rest = select_bin(remaining, live, shrink, TARGET, max_entities=MAX_SITS_PER_PACKAGE)
        if not chosen:
            chosen, rest = [remaining[0]], remaining[1:]
        # fetch + verify; back off the largest slug until the real composed size fits, or it's a lone slug
        xml_text, real = None, 0
        while True:
            print(f"  {pkg_name}: trying {len(chosen)} slugs...", end=" ", flush=True)
            try:
                xml_text = optimise(fetch_bundle(chosen, pkg_name), publisher)
                xml_text = xml_text.replace("\r\n", "\n").replace("\n", "\r\n")
                real = utf16_len(xml_text)
            except urllib.error.HTTPError as e:
                if e.code == 422 and len(chosen) > 1:
                    real = ACCEPT + 1  # force a backoff
                    print("server 422")
                else:
                    print(f"FAILED: {e}")
                    dropped.extend(chosen); chosen = []; break
            except Exception as e:
                # Non-HTTP failure (timeout/DNS/TLS): stay resilient like the measurement phase — drop
                # this bin's slugs (the back-fill pass retries them) instead of aborting the whole build.
                print(f"NETWORK ERROR: {e}")
                dropped.extend(chosen); chosen = []; break
            else:
                print(f"{real//1024}KB UTF-16")
            if real <= ACCEPT or len(chosen) == 1:
                break
            biggest = max(chosen, key=lambda s: live[s])
            chosen.remove(biggest); rest.insert(0, biggest)
            time.sleep(args.delay)
        if not chosen:
            remaining = rest; continue
        # calibrate shrink from this real bin
        ssum = sum(live[s] for s in chosen)
        if ssum: shrink = real / ssum
        if real <= MAX_PACKAGE_UTF16:
            with open(os.path.join(output_dir, f"{pkg_name}.xml"), "w", encoding="utf-8", newline="") as f:
                f.write(xml_text)
            entities = count_entities(xml_text)
            utf8_size = len(xml_text.encode("utf-8"))  # actual on-disk size written
            print(f"  {pkg_name}: OK ({entities} entities, {real//1024}KB UTF-16, {utf8_size//1024}KB UTF-8, {len(chosen)} slugs)")
            results.append({"name": pkg_name, "entities": entities, "size": utf8_size, "utf16bytes": real, "slugs": list(chosen)})
        else:
            print(f"  OVERSIZED single slug '{chosen[0]}' ({real//1024}KB) — exceeds 150KB, cannot deploy as-is")
            dropped.append(chosen[0])
        remaining = rest
        time.sleep(args.delay)
    if remaining:
        dropped.extend(remaining)
        remaining = []

    # ---- Back-fill pass: place any leftover slugs into existing packages that have headroom ----
    # The greedy bins settle below ACCEPT, leaving headroom; small leftovers fit if we add them to
    # an existing package and confirm (fetch-verify) the augmented package still fits. No new pkgs.
    if dropped:
        print(f"  Back-fill pass: {len(dropped)} leftover slug(s), trying existing packages with headroom...")
        still = []
        for slug in dropped:
            placed = False
            # try emptiest (smallest) package first to spread load; skip packages already at entity cap
            for r in sorted(results, key=lambda x: x["size"]):
                if len(r["slugs"]) >= MAX_SITS_PER_PACKAGE:
                    continue
                trial = r["slugs"] + [slug]
                try:
                    txt = optimise(fetch_bundle(trial, r["name"]), publisher)
                    txt = txt.replace("\r\n", "\n").replace("\n", "\r\n")
                    sz = utf16_len(txt)
                except urllib.error.HTTPError as e:
                    if e.code == 422:
                        time.sleep(args.delay); continue
                    print(f"    back-fill fetch failed for {slug} into {r['name']}: {e}"); break
                except Exception as e:
                    print(f"    back-fill network error for {slug} into {r['name']}: {e}"); break
                time.sleep(args.delay)
                if sz <= ACCEPT:
                    with open(os.path.join(output_dir, f"{r['name']}.xml"), "w", encoding="utf-8", newline="") as f:
                        f.write(txt)
                    utf8_sz = len(txt.encode("utf-8"))  # actual on-disk size
                    r["slugs"] = trial; r["size"] = utf8_sz; r["utf16bytes"] = sz; r["entities"] = count_entities(txt)
                    print(f"    + {slug} -> {r['name']} ({sz//1024}KB UTF-16, {utf8_sz//1024}KB UTF-8)")
                    placed = True; break
            if not placed:
                still.append(slug)
        dropped = still
    if dropped:
        print(f"  DROPPED {len(dropped)} slug(s): {dropped}")

    # Write registry
    reg_path = os.path.join(output_dir, "deploy-registry.json")
    with open(reg_path, "w", encoding="utf-8") as f:
        json.dump({
            "tier": args.tier,
            "packages": [{"key": r["name"], "entities": r["entities"], "sizeKB": round(r["size"] / 1024, 1)} for r in results],
        }, f, indent=2)

    total_ents = sum(r["entities"] for r in results)
    print(f"\n  Done: {total_ents} entities across {len(results)} packages")
    if len(results) > MAX_PACKAGES:
        print(f"  WARNING: {len(results)} exceeds {MAX_PACKAGES} package limit!")
    print(f"  Written to {output_dir}")


if __name__ == "__main__":
    main()
