"""
_bfx_loaders.py -- I/O and fetch helpers for Build-FromXLS.py.

Handles spreadsheet loading, TestPattern API fetch, XML package parsing,
and JSON config loading. All pure I/O -- no analysis or reporting.
"""

import json
import os
import re
import sys
import time
import urllib.request
import xml.etree.ElementTree as ET
from collections import defaultdict

GUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", re.I
)

TESTPATTERN_API = "https://testpattern.dev/api/export/purview-bundle"
TESTPATTERN_BATCH_SIZE = 200
TESTPATTERN_PREFIX = "TestPattern - "

# Spreadsheet column indices (0-based) -- matches SIT Risk Analysis sheet
COL_NAME = 0
COL_GUID = 1
COL_CATEGORY = 2
COL_RISK_DESC = 3
COL_RISK_RATING = 4
COL_REF_URL = 5
COL_CLASSIFICATION = 6      # Security classification column (e.g. PSPF level)
COL_FRAMEWORK = 7           # Framework classification (e.g. QGISCF, NZISM)
COL_FRAMEWORK_DLM = 8       # Framework DLM/sublabel
COL_SMALL = 9
COL_MEDIUM = 10
COL_LARGE = 11
COL_LABEL_CODE = 12
COL_CLASSIFIER_TYPE = 13
COL_SOURCE = 14


def _is_y(val):
    return bool(val) and str(val).strip().upper() == "Y"


def load_spreadsheet(xls_path):
    """Load and parse the SIT Risk Analysis sheet from the workbook."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xls_path, read_only=True, data_only=True)

    if "SIT Risk Analysis" not in wb.sheetnames:
        print(f"ERROR: Sheet 'SIT Risk Analysis' not found in {xls_path}", file=sys.stderr)
        print(f"  Available sheets: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb["SIT Risk Analysis"]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[COL_NAME]
        if not name:
            continue
        entries.append({
            "name": str(name).strip(),
            "guid_slug": str(row[COL_GUID] or "").strip(),
            "category": str(row[COL_CATEGORY] or "").strip(),
            "risk_rating": row[COL_RISK_RATING],
            "classification": str(row[COL_CLASSIFICATION] or "").strip(),
            "framework": str(row[COL_FRAMEWORK] or "").strip(),
            "framework_dlm": str(row[COL_FRAMEWORK_DLM] or "").strip(),
            "small": _is_y(row[COL_SMALL]),
            "medium": _is_y(row[COL_MEDIUM]),
            "large": _is_y(row[COL_LARGE]),
            "label_code": str(row[COL_LABEL_CODE] or "").strip(),
            "classifier_type": str(row[COL_CLASSIFIER_TYPE] or "").strip(),
            "source": str(row[COL_SOURCE] or "").strip(),
        })

    wb.close()
    return entries


def load_label_sheet(xls_path, sheet_name="QGISCFDLM"):
    """Load the label definition sheet for label validation."""
    import openpyxl
    wb = openpyxl.load_workbook(xls_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    labels = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            labels.append({
                "name": str(row[0]).strip(),
                "order": row[1],
                "marking": str(row[2] or "").strip(),
                "encrypt": str(row[3] or "").strip(),
                "tooltip": str(row[4] or "").strip(),
            })
    wb.close()
    return labels


def fetch_testpattern_bundles(entries, xml_dir):
    """Fetch XML bundles from testpattern.dev for TestPattern-sourced entries.

    Groups entries by label_code, fetches purview-bundle XML for each group,
    and saves to xml_dir/full/TestPattern_<label>.xml.

    Returns the number of bundles successfully fetched.
    """
    full_dir = os.path.join(xml_dir, "full")
    os.makedirs(full_dir, exist_ok=True)

    # Collect TestPattern entries with slugs (not GUIDs)
    tp_entries = [
        e for e in entries
        if e["source"] == "TestPattern" and not GUID_RE.match(e["guid_slug"])
    ]
    if not tp_entries:
        print("    No TestPattern entries to fetch.")
        return 0

    # Group by label_code
    by_label = defaultdict(list)
    for e in tp_entries:
        by_label[e["label_code"]].append(e["guid_slug"])

    fetched = 0
    errors = []
    headers = {"User-Agent": "Compl8DLPDeploy/1.0"}

    for lc in sorted(by_label.keys()):
        slugs = by_label[lc]
        for i in range(0, len(slugs), TESTPATTERN_BATCH_SIZE):
            batch = slugs[i:i + TESTPATTERN_BATCH_SIZE]
            batch_num = i // TESTPATTERN_BATCH_SIZE + 1
            total_batches = (len(slugs) + TESTPATTERN_BATCH_SIZE - 1) // TESTPATTERN_BATCH_SIZE

            if total_batches > 1:
                bundle_name = f"TestPattern-{lc}-{batch_num}"
                filename = f"TestPattern_{lc}_{batch_num}.xml"
            else:
                bundle_name = f"TestPattern-{lc}"
                filename = f"TestPattern_{lc}.xml"

            slug_param = ",".join(batch)
            url = f"{TESTPATTERN_API}?slugs={slug_param}&name={bundle_name}"
            filepath = os.path.join(full_dir, filename)

            print(f"    {lc} [{batch_num}/{total_batches}] ({len(batch)} patterns)...", end=" ", flush=True)

            try:
                req = urllib.request.Request(url, headers=headers)
                with urllib.request.urlopen(req, timeout=60) as resp:
                    xml_content = resp.read()
                    with open(filepath, "wb") as f:
                        f.write(xml_content)
                    print(f"OK -> {filename} ({len(xml_content):,} bytes)")
                    fetched += 1
            except Exception as e:
                print(f"FAILED: {e}")
                errors.append((lc, batch_num, str(e)))

            time.sleep(0.5)

    if errors:
        print(f"    WARNING: {len(errors)} fetch errors:")
        for lc, bn, err in errors:
            print(f"      {lc} batch {bn}: {err}")

    return fetched


def build_xml_guid_lookup(xml_dir):
    """Parse all XML rule packages to build name->GUID lookup.

    Handles the "TestPattern - " prefix convention: entries named
    "TestPattern - Foo" are indexed under both the prefixed and
    unprefixed ("Foo") forms for matching against spreadsheet names.
    """
    lookup = {}  # lowercase name -> {guid, name, package}
    full_dir = os.path.join(xml_dir, "full")
    if not os.path.isdir(full_dir):
        print(f"  WARNING: XML directory not found: {full_dir}", file=sys.stderr)
        return lookup

    for fname in sorted(os.listdir(full_dir)):
        if not fname.endswith(".xml"):
            continue
        path = os.path.join(full_dir, fname)
        try:
            tree = ET.parse(path)
            root = tree.getroot()
        except ET.ParseError as e:
            print(f"  WARNING: Failed to parse {fname}: {e}", file=sys.stderr)
            continue

        # Build Resource id->name map
        resources = {}
        for elem in root.iter():
            if "Resource" in elem.tag:
                id_ref = elem.get("idRef", "")
                for child in elem:
                    if "Name" in child.tag:
                        text = child.text or child.get("default", "")
                        if text:
                            resources[id_ref] = text.strip()
                        break

        # Find Entity elements
        for elem in root.iter():
            if "Entity" in elem.tag and elem.get("id"):
                eid = elem.get("id")
                ename = resources.get(eid, "")
                if ename:
                    entry = {"guid": eid, "name": ename, "package": fname}
                    lookup[ename.lower()] = entry
                    # Also index without "TestPattern - " prefix
                    if ename.lower().startswith(TESTPATTERN_PREFIX.lower()):
                        unprefixed = ename[len(TESTPATTERN_PREFIX):]
                        lookup[unprefixed.lower()] = entry
    return lookup


def build_existing_classifier_lookup(config_dir):
    """Load current classifiers.json to use as fallback GUID source."""
    path = os.path.join(config_dir, "classifiers.json")
    if not os.path.isfile(path):
        return {}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    lookup = {}  # lowercase name -> {id, classifierType}
    for _label_code, items in data.items():
        for item in items:
            name_lower = item["name"].lower()
            if name_lower not in lookup:
                lookup[name_lower] = {
                    "id": item.get("id", ""),
                    "classifierType": item.get("classifierType", ""),
                }
    return lookup


def load_labels_json(config_dir):
    """Load labels.json for validation."""
    path = os.path.join(config_dir, "labels.json")
    if not os.path.isfile(path):
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)
