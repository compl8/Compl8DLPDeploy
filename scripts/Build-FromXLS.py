#!/usr/bin/env python3
"""
Build-FromXLS.py -- Generates DLP config from SIT Risk Analysis spreadsheet.

Reads the SIT Risk Analysis Excel workbook and generates:
  - config/classifiers.json (classifier -> label code mapping for DLP rules)
  - config/labels.json updates (validates label definitions match label sheet)
  - Gap report (classifiers needing XML packages, GUID resolution issues)

Automatically fetches TestPattern XML bundles from testpattern.dev API for any
TestPattern-sourced entries that lack GUIDs. Use --skip-fetch to disable.

Usage:
  python scripts/Build-FromXLS.py                                    # defaults
  python scripts/Build-FromXLS.py --xls SIT-Risk-Analysis-v11.xlsx   # specify file
  python scripts/Build-FromXLS.py --tier large                       # deployment tier
  python scripts/Build-FromXLS.py --dry-run                          # report only, no writes
  python scripts/Build-FromXLS.py --skip-fetch                       # skip testpattern.dev fetch
  python scripts/Build-FromXLS.py --label-sheet QGISCFDLM            # specify label sheet name
"""

import argparse
import json
import os
import re
import sys
import time
import urllib.request
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path

GUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", re.I
)

TESTPATTERN_API = "https://testpattern.dev/api/export/purview-bundle"
TESTPATTERN_BATCH_SIZE = 200
TESTPATTERN_PREFIX = "TestPattern - "

# Spreadsheet column indices (0-based) -- matches SIT Risk Analysis sheet
COL_NAME = 0
COL_GUID = 1
COL_CATEGORY = 2
COL_RISK_DESC = 3
COL_RISK_RATING = 4
COL_REF_URL = 5
COL_CLASSIFICATION = 6      # Security classification column (e.g. PSPF level)
COL_FRAMEWORK = 7           # Framework classification (e.g. QGISCF, NZISM)
COL_FRAMEWORK_DLM = 8       # Framework DLM/sublabel
COL_SMALL = 9
COL_MEDIUM = 10
COL_LARGE = 11
COL_LABEL_CODE = 12
COL_CLASSIFIER_TYPE = 13
COL_SOURCE = 14


def load_spreadsheet(xls_path):
    """Load and parse the SIT Risk Analysis sheet from the workbook."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xls_path, read_only=True, data_only=True)

    if "SIT Risk Analysis" not in wb.sheetnames:
        print(f"ERROR: Sheet 'SIT Risk Analysis' not found in {xls_path}", file=sys.stderr)
        print(f"  Available sheets: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb["SIT Risk Analysis"]
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[COL_NAME]
        if not name:
            continue
        entries.append({
            "name": str(name).strip(),
            "guid_slug": str(row[COL_GUID] or "").strip(),
            "category": str(row[COL_CATEGORY] or "").strip(),
            "risk_rating": row[COL_RISK_RATING],
            "classification": str(row[COL_CLASSIFICATION] or "").strip(),
            "framework": str(row[COL_FRAMEWORK] or "").strip(),
            "framework_dlm": str(row[COL_FRAMEWORK_DLM] or "").strip(),
            "small": _is_y(row[COL_SMALL]),
            "medium": _is_y(row[COL_MEDIUM]),
            "large": _is_y(row[COL_LARGE]),
            "label_code": str(row[COL_LABEL_CODE] or "").strip(),
            "classifier_type": str(row[COL_CLASSIFIER_TYPE] or "").strip(),
            "source": str(row[COL_SOURCE] or "").strip(),
        })

    wb.close()
    return entries


def load_label_sheet(xls_path, sheet_name="QGISCFDLM"):
    """Load the label definition sheet for label validation."""
    import openpyxl
    wb = openpyxl.load_workbook(xls_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    labels = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            labels.append({
                "name": str(row[0]).strip(),
                "order": row[1],
                "marking": str(row[2] or "").strip(),
                "encrypt": str(row[3] or "").strip(),
                "tooltip": str(row[4] or "").strip(),
            })
    wb.close()
    return labels


def _is_y(val):
    return bool(val) and str(val).strip().upper() == "Y"


def fetch_testpattern_bundles(entries, xml_dir):
    """Fetch XML bundles from testpattern.dev for TestPattern-sourced entries.

    Groups entries by label_code, fetches purview-bundle XML for each group,
    and saves to xml_dir/full/TestPattern_<label>.xml.

    Returns the number of bundles successfully fetched.
    """
    full_dir = os.path.join(xml_dir, "full")
    os.makedirs(full_dir, exist_ok=True)

    # Collect TestPattern entries with slugs (not GUIDs)
    tp_entries = [
        e for e in entries
        if e["source"] == "TestPattern" and not GUID_RE.match(e["guid_slug"])
    ]
    if not tp_entries:
        print("    No TestPattern entries to fetch.")
        return 0

    # Group by label_code
    by_label = defaultdict(list)
    for e in tp_entries:
        by_label[e["label_code"]].append(e["guid_slug"])

    fetched = 0
    errors = []
    headers = {"User-Agent": "Compl8DLPDeploy/1.0"}

    for lc in sorted(by_label.keys()):
        slugs = by_label[lc]
        for i in range(0, len(slugs), TESTPATTERN_BATCH_SIZE):
            batch = slugs[i:i + TESTPATTERN_BATCH_SIZE]
            batch_num = i // TESTPATTERN_BATCH_SIZE + 1
            total_batches = (len(slugs) + TESTPATTERN_BATCH_SIZE - 1) // TESTPATTERN_BATCH_SIZE

            if total_batches > 1:
                bundle_name = f"TestPattern-{lc}-{batch_num}"
                filename = f"TestPattern_{lc}_{batch_num}.xml"
            else:
                bundle_name = f"TestPattern-{lc}"
                filename = f"TestPattern_{lc}.xml"

            slug_param = ",".join(batch)
            url = f"{TESTPATTERN_API}?slugs={slug_param}&name={bundle_name}"
            filepath = os.path.join(full_dir, filename)

            print(f"    {lc} [{batch_num}/{total_batches}] ({len(batch)} patterns)...", end=" ", flush=True)

            try:
                req = urllib.request.Request(url, headers=headers)
                with urllib.request.urlopen(req, timeout=60) as resp:
                    xml_content = resp.read()
                    with open(filepath, "wb") as f:
                        f.write(xml_content)
                    print(f"OK -> {filename} ({len(xml_content):,} bytes)")
                    fetched += 1
            except Exception as e:
                print(f"FAILED: {e}")
                errors.append((lc, batch_num, str(e)))

            time.sleep(0.5)

    if errors:
        print(f"    WARNING: {len(errors)} fetch errors:")
        for lc, bn, err in errors:
            print(f"      {lc} batch {bn}: {err}")

    return fetched


def build_xml_guid_lookup(xml_dir):
    """Parse all XML rule packages to build name->GUID lookup.

    Handles the "TestPattern - " prefix convention: entries named
    "TestPattern - Foo" are indexed under both the prefixed and
    unprefixed ("Foo") forms for matching against spreadsheet names.
    """
    lookup = {}  # lowercase name -> {guid, name, package}
    full_dir = os.path.join(xml_dir, "full")
    if not os.path.isdir(full_dir):
        print(f"  WARNING: XML directory not found: {full_dir}", file=sys.stderr)
        return lookup

    for fname in sorted(os.listdir(full_dir)):
        if not fname.endswith(".xml"):
            continue
        path = os.path.join(full_dir, fname)
        try:
            tree = ET.parse(path)
            root = tree.getroot()
        except ET.ParseError as e:
            print(f"  WARNING: Failed to parse {fname}: {e}", file=sys.stderr)
            continue

        # Build Resource id->name map
        resources = {}
        for elem in root.iter():
            if "Resource" in elem.tag:
                id_ref = elem.get("idRef", "")
                for child in elem:
                    if "Name" in child.tag:
                        text = child.text or child.get("default", "")
                        if text:
                            resources[id_ref] = text.strip()
                        break

        # Find Entity elements
        for elem in root.iter():
            if "Entity" in elem.tag and elem.get("id"):
                eid = elem.get("id")
                ename = resources.get(eid, "")
                if ename:
                    entry = {"guid": eid, "name": ename, "package": fname}
                    lookup[ename.lower()] = entry
                    # Also index without "TestPattern - " prefix
                    if ename.lower().startswith(TESTPATTERN_PREFIX.lower()):
                        unprefixed = ename[len(TESTPATTERN_PREFIX):]
                        lookup[unprefixed.lower()] = entry
    return lookup


def build_existing_classifier_lookup(config_dir):
    """Load current classifiers.json to use as fallback GUID source."""
    path = os.path.join(config_dir, "classifiers.json")
    if not os.path.isfile(path):
        return {}
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    lookup = {}  # lowercase name -> {id, classifierType}
    for _label_code, items in data.items():
        for item in items:
            name_lower = item["name"].lower()
            if name_lower not in lookup:
                lookup[name_lower] = {
                    "id": item.get("id", ""),
                    "classifierType": item.get("classifierType", ""),
                }
    return lookup



def load_labels_json(config_dir):
    """Load labels.json for validation."""
    path = os.path.join(config_dir, "labels.json")
    if not os.path.isfile(path):
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def resolve_guid(entry, xml_lookup, classifier_lookup):
    """Resolve GUID for an entry using available sources.

    Returns (guid, source) where source is one of:
      'spreadsheet' - GUID from XLS GUID column
      'xml' - GUID from XML package entity
      'existing' - GUID from current classifiers.json
      'unresolved' - no GUID found
    """
    # 1. Spreadsheet GUID column (actual UUID format)
    if GUID_RE.match(entry["guid_slug"]):
        return entry["guid_slug"], "spreadsheet"

    # 2. XML package entity lookup (exact name match, including prefix-stripped)
    name_lower = entry["name"].lower()
    if name_lower in xml_lookup:
        return xml_lookup[name_lower]["guid"], "xml"

    # 3. Current classifiers.json fallback
    if name_lower in classifier_lookup:
        existing = classifier_lookup[name_lower]
        if existing["id"]:
            return existing["id"], "existing"

    return entry["guid_slug"], "unresolved"


def filter_by_tier(entries, tier):
    """Filter entries by deployment tier (small/medium/large)."""
    if tier == "large":
        return [e for e in entries if e["large"]]
    elif tier == "medium":
        return [e for e in entries if e["medium"]]
    elif tier == "small":
        return [e for e in entries if e["small"]]
    else:
        raise ValueError(f"Unknown tier: {tier}")


def build_classifier_entry(entry, guid, classifier_type):
    """Build a classifiers.json entry for the given classifier."""
    if classifier_type == "MLModel":
        return {
            "name": entry["name"],
            "id": guid,
            "classifierType": "MLModel",
        }
    else:
        # SIT, Regex, Keyword List all become SIT entries in DLP rules
        result = {
            "name": entry["name"],
            "id": guid,
        }
        # Add minCount/maxCount/confidenceLevel overrides if specified
        # (defaults are applied by Resolve-ClassifierConfig in the module)
        if entry.get("minCount"):
            result["minCount"] = entry["minCount"]
        if entry.get("maxCount"):
            result["maxCount"] = entry["maxCount"]
        if entry.get("confidenceLevel"):
            result["confidenceLevel"] = entry["confidenceLevel"]
        return result


def build_classifiers_json(entries, xml_lookup, classifier_lookup):
    """Build the full classifiers.json structure from spreadsheet entries.

    Returns (classifiers_dict, stats_dict).
    """
    classifiers = defaultdict(list)
    stats = {
        "resolved": {"spreadsheet": 0, "xml": 0, "existing": 0},
        "unresolved": [],
        "by_label": defaultdict(int),
        "by_type": defaultdict(int),
        "by_source": defaultdict(int),
    }

    # Deduplicate by (label_code, name)
    seen = set()

    for entry in entries:
        label_code = entry["label_code"]
        if not label_code:
            continue

        key = (label_code, entry["name"])
        if key in seen:
            continue
        seen.add(key)

        guid, source = resolve_guid(entry, xml_lookup, classifier_lookup)

        # Map classifier type from spreadsheet to classifiers.json type
        ct = entry["classifier_type"]
        if ct == "MLModel":
            cfg_type = "MLModel"
        else:
            cfg_type = "SIT"  # SIT, Regex, Keyword List all become SIT entries

        entry_obj = build_classifier_entry(entry, guid, cfg_type)
        classifiers[label_code].append(entry_obj)

        stats["by_label"][label_code] += 1
        stats["by_type"][ct] += 1
        stats["by_source"][entry["source"]] += 1

        if source == "unresolved":
            stats["unresolved"].append({
                "name": entry["name"],
                "label_code": label_code,
                "type": ct,
                "source": entry["source"],
                "slug": entry["guid_slug"],
            })
        else:
            stats["resolved"][source] += 1

    # Sort entries within each label code by name
    for label_code in classifiers:
        classifiers[label_code] = sorted(
            classifiers[label_code], key=lambda x: x["name"]
        )

    # Sort label codes alphabetically
    ordered = {}
    for k in sorted(classifiers.keys()):
        ordered[k] = classifiers[k]

    return ordered, stats


def validate_labels(labels_json, label_sheet_data, sheet_name="QGISCFDLM"):
    """Compare labels.json against label sheet for discrepancies."""
    issues = []

    # Build lookup from labels.json
    json_labels = {}
    for lbl in labels_json:
        display = lbl.get("displayName", "")
        if display:
            json_labels[display] = lbl

    # Check each label sheet entry
    for xl in label_sheet_data:
        name = xl["name"]
        if name not in json_labels:
            issues.append(f"{sheet_name} label '{name}' not found in labels.json")
            continue

        jl = json_labels[name]

        # Check tooltip text matches
        xl_tip = xl["tooltip"].replace(" -- ", " \u2014 ") if xl["tooltip"] else ""
        jl_tip = jl.get("tooltip", "")
        if xl_tip and jl_tip and not jl_tip.startswith(xl_tip[:50]):
            issues.append(f"Tooltip mismatch for '{name}': XLS starts '{xl_tip[:60]}...', JSON starts '{jl_tip[:60]}...'")

        # Check visual marking
        if xl["marking"] and jl.get("headerText"):
            if xl["marking"] != jl["headerText"]:
                issues.append(f"Header marking mismatch for '{name}': XLS='{xl['marking']}', JSON='{jl['headerText']}'")

    # Check for labels.json entries not in label sheet
    xl_names = {lbl["name"] for lbl in label_sheet_data}
    for jl in labels_json:
        dn = jl.get("displayName", "")
        if dn and not jl.get("isGroup") and dn not in xl_names:
            if jl.get("code") and "IT" in jl.get("code", ""):
                issues.append(f"labels.json has '{dn}' (code: {jl['code']}) but {sheet_name} sheet does not -- consider adding InfoTech labels to spreadsheet")
            else:
                issues.append(f"labels.json has '{dn}' but not in {sheet_name} sheet")

    return issues


def print_report(stats, label_issues, total_entries, tier):
    """Print the gap analysis report."""
    print(f"\n{'='*70}")
    print(f"  DLP Build Report -- Tier: {tier}")
    print(f"{'='*70}")

    resolved_total = sum(stats["resolved"].values())
    unresolved_total = len(stats["unresolved"])
    print(f"\n  Entries processed: {total_entries}")
    print(f"  GUIDs resolved:   {resolved_total}")
    print(f"    From spreadsheet: {stats['resolved']['spreadsheet']}")
    print(f"    From XML packages: {stats['resolved']['xml']}")
    print(f"    From existing config: {stats['resolved']['existing']}")
    print(f"  Unresolved (need XML): {unresolved_total}")

    print(f"\n  By label code:")
    for lc in sorted(stats["by_label"]):
        print(f"    {lc:12s}: {stats['by_label'][lc]:4d}")

    print(f"\n  By classifier type:")
    for ct in sorted(stats["by_type"]):
        print(f"    {ct:15s}: {stats['by_type'][ct]:4d}")

    print(f"\n  By source:")
    for src in sorted(stats["by_source"]):
        print(f"    {src:20s}: {stats['by_source'][src]:4d}")

    if stats["unresolved"]:
        print(f"\n  --- Unresolved Entries ({unresolved_total}) ---")
        print(f"  These classifiers need XML SIT packages before DLP deployment:")
        by_lc = defaultdict(list)
        for u in stats["unresolved"]:
            by_lc[u["label_code"]].append(u)
        for lc in sorted(by_lc):
            items = by_lc[lc]
            print(f"\n    [{lc}] ({len(items)} entries)")
            for item in sorted(items, key=lambda x: x["name"])[:10]:
                print(f"      {item['name']} ({item['type']}, {item['source']})")
            if len(items) > 10:
                print(f"      ... and {len(items) - 10} more")

    if label_issues:
        print(f"\n  --- Label Validation Issues ({len(label_issues)}) ---")
        for issue in label_issues:
            print(f"    - {issue}")

    print(f"\n{'='*70}")


def main():
    parser = argparse.ArgumentParser(
        description="Build DLP config from SIT Risk Analysis spreadsheet"
    )
    parser.add_argument(
        "--xls",
        default=None,
        help="Path to SIT Risk Analysis Excel file (default: auto-detect in project root)",
    )
    parser.add_argument(
        "--config-dir",
        default="config",
        help="Path to config directory (default: config/)",
    )
    parser.add_argument(
        "--xml-dir",
        default="xml",
        help="Path to XML packages directory (default: xml/)",
    )
    parser.add_argument(
        "--tier",
        choices=["small", "medium", "large"],
        default="large",
        help="Deployment tier to build for (default: large)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report only, don't write config files",
    )
    parser.add_argument(
        "--skip-fetch",
        action="store_true",
        help="Skip fetching TestPattern XML bundles from testpattern.dev",
    )
    parser.add_argument(
        "--include-unresolved",
        action="store_true",
        help="Include unresolved entries (slug as ID) in classifiers.json",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output path for classifiers.json (default: config/classifiers.json)",
    )
    parser.add_argument(
        "--label-sheet",
        default="QGISCFDLM",
        help="Name of the label definition sheet in the workbook (default: QGISCFDLM)",
    )

    args = parser.parse_args()

    # Resolve project root
    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parent
    config_dir = (project_root / args.config_dir).resolve()
    xml_dir = (project_root / args.xml_dir).resolve()

    # Auto-detect XLS file
    if args.xls:
        xls_path = Path(args.xls).resolve()
    else:
        candidates = sorted(project_root.glob("SIT-Risk-Analysis-v*.xlsx"), reverse=True)
        if candidates:
            xls_path = candidates[0]
        else:
            print("ERROR: No SIT-Risk-Analysis-v*.xlsx found in project root.", file=sys.stderr)
            print("  Specify with --xls <path>", file=sys.stderr)
            sys.exit(1)

    print(f"  Source XLS:     {xls_path.name}")
    print(f"  Config dir:     {config_dir}")
    print(f"  XML dir:        {xml_dir}")
    print(f"  Tier:           {args.tier}")
    print(f"  TestPattern:    {'skip' if args.skip_fetch else 'fetch from testpattern.dev'}")

    # Load data sources
    print(f"\n  Loading spreadsheet...")
    all_entries = load_spreadsheet(str(xls_path))
    print(f"    {len(all_entries)} total entries")

    entries = filter_by_tier(all_entries, args.tier)
    print(f"    {len(entries)} entries for tier '{args.tier}'")

    # Fetch TestPattern XML bundles
    if not args.skip_fetch:
        tp_count = sum(1 for e in entries if e["source"] == "TestPattern" and not GUID_RE.match(e["guid_slug"]))
        if tp_count > 0:
            print(f"\n  Fetching TestPattern XML bundles ({tp_count} entries)...")
            fetched = fetch_testpattern_bundles(entries, str(xml_dir))
            print(f"    {fetched} bundles fetched")

    print(f"\n  Loading XML packages...")
    xml_lookup = build_xml_guid_lookup(str(xml_dir))
    print(f"    {len(xml_lookup)} entities from XML")

    print(f"  Loading existing classifiers.json...")
    classifier_lookup = build_existing_classifier_lookup(str(config_dir))
    print(f"    {len(classifier_lookup)} existing entries")

    # Load label definitions for validation
    label_sheet_data = load_label_sheet(str(xls_path), args.label_sheet)
    labels_json = load_labels_json(str(config_dir))
    label_issues = validate_labels(labels_json, label_sheet_data, args.label_sheet)

    # Build classifiers.json
    classifiers, stats = build_classifiers_json(
        entries, xml_lookup, classifier_lookup,
    )

    # Strip unresolved entries unless --include-unresolved
    if not args.include_unresolved:
        unresolved_names = {u["name"] for u in stats["unresolved"]}
        for label_code in list(classifiers.keys()):
            filtered = [e for e in classifiers[label_code] if e["name"] not in unresolved_names]
            if filtered:
                classifiers[label_code] = filtered
            else:
                del classifiers[label_code]

    # Print report
    print_report(stats, label_issues, len(entries), args.tier)

    # Count final output
    total_entries_out = sum(len(v) for v in classifiers.values())
    unique_names_out = len({e["name"] for v in classifiers.values() for e in v})
    print(f"\n  Output classifiers.json:")
    print(f"    {len(classifiers)} label codes")
    print(f"    {total_entries_out} total entries ({unique_names_out} unique classifiers)")
    for lc in classifiers:
        print(f"      {lc:12s}: {len(classifiers[lc]):4d}")

    if args.dry_run:
        print(f"\n  [DRY RUN] No files written.")
        return

    # Write classifiers.json
    output_path = Path(args.output) if args.output else config_dir / "classifiers.json"
    # Back up existing file
    if output_path.exists():
        import shutil
        backup_path = output_path.with_suffix(".json.bak")
        print(f"\n  Backing up existing: {output_path.name} -> {backup_path.name}")
        shutil.copy2(str(output_path), str(backup_path))

    with open(str(output_path), "w", encoding="utf-8") as f:
        json.dump(classifiers, f, indent=4, ensure_ascii=False)
    print(f"  Written: {output_path}")

    # Write unresolved report
    if stats["unresolved"]:
        report_path = config_dir / "classifiers-unresolved.json"
        with open(str(report_path), "w", encoding="utf-8") as f:
            json.dump(stats["unresolved"], f, indent=4, ensure_ascii=False)
        print(f"  Written: {report_path} ({len(stats['unresolved'])} entries needing XML)")
    else:
        # Clean up stale unresolved file
        report_path = config_dir / "classifiers-unresolved.json"
        if report_path.exists():
            report_path.unlink()
            print(f"  Removed: {report_path} (all entries resolved)")

    print(f"\n  Done.")


if __name__ == "__main__":
    main()
